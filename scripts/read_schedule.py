#!/usr/bin/env python3
"""Dump the 연찬회 schedule xlsx so we can build the real itinerary."""
import openpyxl
from pathlib import Path

src = Path.home() / "Downloads" / "도시과학대학연찬회 세부일정.xlsx"
out = Path(__file__).resolve().parent.parent / "docs" / "_schedule_dump.txt"
wb = openpyxl.load_workbook(src, data_only=True)
lines = []
for ws in wb.worksheets:
    lines.append(f"\n===== SHEET: {ws.title}  ({ws.max_row} rows x {ws.max_column} cols) =====")
    for row in ws.iter_rows(values_only=True):
        cells = [("" if c is None else str(c)).strip() for c in row]
        if any(cells):
            # collapse trailing empties for readability
            while cells and cells[-1] == "":
                cells.pop()
            lines.append(" | ".join(cells))
out.write_text("\n".join(lines), encoding="utf-8")
print(f"wrote {out} ({len(lines)} lines)")
